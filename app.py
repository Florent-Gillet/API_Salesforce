from flask import Flask, request, jsonify
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/')
def home():
    return 'API de décodage prête.'

@app.route('/decode', methods=['POST'])
def decode():
    data = request.get_json()
    print("Data reçu:", data)  # Ajout de log pour vérifier si on reçoit des données

    code = data.get('code')

    # Charger le fichier Excel (avec les valeurs déjà remplies)
    wb = load_workbook('decoder.xlsx', data_only=True)
    ws = wb['DECODER']

    # Lire les valeurs dans C100:C114 (produits) et D100:D114 (quantités)
    produits = []
    for i in range(100, 115):
        produit = ws[f'C{i}'].value
        quantite = ws[f'D{i}'].value
        if produit and quantite:
            produits.append({
                'product_name': produit,
                'quantity': quantite
            })
    
    print("Réponse produits:", produits)  # Log de la réponse

    return jsonify(produits)

if __name__ == '__main__':
    app.run(debug=True)  # Ajout de `debug=True` pour forcer le mode debug
